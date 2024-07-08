'''
FLEXICAL v3 | TEMPLATES

'''

## PYTHON LIBRARIES
import os
from datetime import datetime
from typing import Dict
from dataclasses import asdict
from enum import Enum

## IMPORTED LIBRARIES
import streamlit as st
import pandas as pd

## INTERNAL
from app import *
from db import *

import func_xlsx as XLS



## SESSION STATES
## __________________________________________________________________________________________________

if 'LOGIN_STATUS' not in st.session_state:
    st.session_state.LOGIN_STATUS = None

if 'TEMPLATES' not in st.session_state:
    st.session_state.TEMPLATES = 1

class TEMPLATE:

    class TypeDict(TypedDict):
        Id: str
        MODEL_ID: str
        VARSION: str
        INFO: str
        DB: dict
        PYDATA: str

    @dataclass
    class DB:
        TEST_LIST: dict = None
        def toDict(self) -> Dict[str, any]:
            return asdict(self)
    
    @dataclass
    class TEST:
        TEST: str
        PARAMETERS: str = None
        CONFIG: str = None
        INFO: str = None
        PROCEDURE_ID: str = None
        CALIBRATION: dict = None
        def toDict(self) -> Dict[str, any]:
            return asdict(self)

    class MEASURE(Enum):
        RANGE_TX = 0
        RANGE = auto()
        VALUE1 = auto()
        VALUE2 = auto()
        # MEASURE = auto()
        # DEVIATION = auto()
        # SPECIFICATION = auto() # LIMIT OF ERROR
        # UNCERTAINTY = auto()
        # RESULT = auto()
        # CMC = auto()
        # ACQUISITIONS = auto() # Podria estar dentro de VALIDATION
        # VALIDATION = auto() # K_FACTOR, PROCEDURE, STANDARDS

## MENU
## __________________________________________________________________________________________________

def TEST_EDITOR(ID: str, DB: dict) -> None:
    procedures = SQL_SELECT_COLUMN("PROCEDURES", "Id")

    @st.experimental_dialog(title='✏️ EDITOR', width='small')
    def FORM_NEW():
        
        PROCEDURE_ID = st.selectbox("PROCEDURE Id *", options=procedures, index=None)
        title = str()
        if PROCEDURE_ID:
            title = SQL_BY_ROW("PROCEDURES", "Id", PROCEDURE_ID)[0]['TITLE']
        TEST = st.text_input("TEST TITLE *", value=title)
        PARAMETERS = st.text_input("TEST PARAMETERS")
        CONFIG = st.text_input("CONFIG & CONNECTIONS")
        INFO = st.text_area("INFO")
        if st.button("➕ INSERT NEW TEST", key='btn_test_ADD'):
            NEW_TEST = TEMPLATE.TEST(TEST, PARAMETERS, CONFIG, INFO, PROCEDURE_ID, CALIBRATION={})
            DB["TEST_LIST"].append(NEW_TEST.toDict())
            try:
                SQL_UPDATE_DB("TEMPLATES", ID, DB)
                st.session_state.TEMPLATES += 1
                st.rerun()
            except Exception as e:
                INFOBOX(e)

    @st.experimental_dialog(title='✏️ EDITOR', width='small')
    def FORM_EDIT(TEST: TEMPLATE.TEST, loc):
        indx = None
        if TEST.PROCEDURE_ID in procedures:
            indx = procedures.index(TEST.PROCEDURE_ID)
        PROCEDURE_ID = st.selectbox("PROCEDURE Id", options=procedures, index=indx)
        TITLE = st.text_input("TEST TITLE *", value=TEST.TEST)
        PARAMETERS = st.text_input("TEST PARAMETERS", value=TEST.PARAMETERS)
        CONFIG = st.text_input("CONFIG & CONNECTIONS", value=TEST.CONFIG)
        INFO = st.text_area("INFO", value=TEST.INFO)
        if st.button(USUAL_ICONS.UPDATE.value + " UPDATE", key='btn_test_update'):
            EDIT_TEST = TEMPLATE.TEST(TITLE, PARAMETERS, CONFIG, INFO, PROCEDURE_ID, TEST.CALIBRATION)
            DB["TEST_LIST"][loc] = EDIT_TEST.toDict()
            try:
                SQL_UPDATE_DB("TEMPLATES", ID, DB)
                st.session_state.TEMPLATES += 1
                st.rerun()
            except Exception as e:
                INFOBOX(e)

    ## TABLE TEST
    DATAFRAME = pd.DataFrame(DB['TEST_LIST'], columns=['TEST', 'PARAMETERS', 'CONFIG', 'INFO', 'PROCEDURE_ID'])

    col12, col22 = st.columns([9,1])

    with col12:
        TBL_TEST = st.dataframe(
            data=DATAFRAME, 
            hide_index=True,
            on_select="rerun", # Con esta opcion aparece el selector de fila
            selection_mode=['single-row'], # "multi-column" "multi-row"
            use_container_width=True,
        )
    
    with col22:
        with st.popover(USUAL_ICONS.EXPANDER.value):
            if st.button(label='➕ ADD TEST', use_container_width=True):
                FORM_NEW()

            if len(TBL_TEST.selection.rows) == 1:
                loc = TBL_TEST.selection.rows[0]
                test = TEMPLATE.TEST(**DB["TEST_LIST"][loc])

                if st.button(label='✏️ EDIT TEST', use_container_width=True):
                    FORM_EDIT(test, loc)
                
                if st.button(label='➖ DEL TEST', use_container_width=True):
                    del DB['TEST_LIST'][loc]
                    try:
                        SQL_UPDATE_DB("TEMPLATES", ID, DB)
                        st.session_state.TEMPLATES += 1
                        st.rerun()
                    except Exception as e:
                        INFOBOX(e)
                
                if loc > 0:
                    if st.button(label=USUAL_ICONS.UP.value, use_container_width=True):
                        DB['TEST_LIST'].insert(loc-1, DB['TEST_LIST'].pop(loc))
                        SQL_UPDATE_DB("TEMPLATES", ID, DB)
                        st.session_state.TEMPLATES += 1
                        st.rerun()
                
                if loc < len(DB['TEST_LIST'])-1:
                    if st.button(label=USUAL_ICONS.DOWN.value, use_container_width=True):
                        DB['TEST_LIST'].insert(loc+1, DB['TEST_LIST'].pop(loc))
                        SQL_UPDATE_DB("TEMPLATES", ID, DB)
                        st.session_state.TEMPLATES += 1
                        st.rerun()

    if len(TBL_TEST.selection.rows) == 1:
        st.text("")
        DF_CALIBRATION = pd.DataFrame(test.CALIBRATION, columns=[field.name for field in TEMPLATE.MEASURE])
        DF_CALIBRATION['RANGE_TX'] = DF_CALIBRATION['RANGE_TX'].astype(str)
        DF_CALIBRATION['RANGE'] = DF_CALIBRATION['RANGE'].astype(float)
        DF_CALIBRATION['VALUE1'] = DF_CALIBRATION['VALUE1'].astype(float)
        DF_CALIBRATION['VALUE2'] = DF_CALIBRATION['VALUE2'].astype(float)
        DF_CALIBRATION = DF_CALIBRATION.reset_index()
        del DF_CALIBRATION['index']
        tbl_cal_test = st.data_editor(
            # data=pd.DataFrame(test.CALIBRATION, columns=[field.name for field in TEMPLATE.MEASURE]),
            data=DF_CALIBRATION,
            use_container_width=True,
            hide_index=True,
            column_config={
                'RANGE_TX': st.column_config.TextColumn(default=""),
            }, 
            num_rows='dynamic'
        )
        if st.button(USUAL_ICONS.UPDATE.value + " UPDATE", key='btn_tbl_update'):
            DB["TEST_LIST"][loc]['CALIBRATION'] = tbl_cal_test.to_dict()
            try:
                SQL_UPDATE_DB("TEMPLATES", ID, DB)
                st.session_state.TEMPLATES += 1
                st.rerun()
            except Exception as e:
                INFOBOX(e)


from openpyxl.styles import PatternFill, Protection
from openpyxl.formatting.rule import FormulaRule


def PRINT_XLSX(CURRENT_TEMPLATE: TEMPLATE.TypeDict) -> None:

    class HEADERS_NOVISIBLE(Enum):
        ACQ_MEASURE = 9
        ACQ_INDICATION = 20
        PROCEDURE_ID = 32
        MEAS_TYPE = 33
        NOT_INSTALLED = 34
        RESULT = 35
        RANGE_TX = 38
        RANGE = 40
        VALUE1 = 42
        VALUE2 = 44
        INDICATION = 46
        MEASURE = 48
        DEVIATION = 50
        SPECIFICATION = 52
        CMC = 54
        UNCERTAINTY = 56
        U = 58
        K = 60
        RESOLUTION = 62
        TYPB = 64
        TYPA_MEAS = 66
        TYPA_INDI = 68

    def EXCEL_FORMULA_RESULT(ROW: int):
        FORMULA: str = \
f'''=
IF(NOT(ISBLANK({XLS.CELL(ROW, HEADERS_NOVISIBLE.NOT_INSTALLED.value)})),"n.i.",
IF({XLS.CELL(ROW, HEADERS_NOVISIBLE.MEASURE.value)}=FALSE,"n.m.",
IF({XLS.CELL(ROW, HEADERS_NOVISIBLE.MEAS_TYPE.value)}="DEVIATION",
    IF(ABS({XLS.CELL(ROW, HEADERS_NOVISIBLE.DEVIATION.value)})+ABS({XLS.CELL(ROW, HEADERS_NOVISIBLE.UNCERTAINTY.value)})<ABS({XLS.CELL(ROW, HEADERS_NOVISIBLE.SPECIFICATION.value)}),"1 Pass",
    IF(ABS({XLS.CELL(ROW, HEADERS_NOVISIBLE.DEVIATION.value)})+ABS({XLS.CELL(ROW, HEADERS_NOVISIBLE.UNCERTAINTY.value)})=ABS({XLS.CELL(ROW, HEADERS_NOVISIBLE.SPECIFICATION.value)}),"2 Pass",
    IF(AND(ABS({XLS.CELL(ROW, HEADERS_NOVISIBLE.DEVIATION.value)})<ABS({XLS.CELL(ROW, HEADERS_NOVISIBLE.SPECIFICATION.value)}),ABS({XLS.CELL(ROW, HEADERS_NOVISIBLE.DEVIATION.value)})+ABS({XLS.CELL(ROW, HEADERS_NOVISIBLE.UNCERTAINTY.value)})>ABS({XLS.CELL(ROW, HEADERS_NOVISIBLE.SPECIFICATION.value)})),"3 UGB1",
    IF(ABS({XLS.CELL(ROW, HEADERS_NOVISIBLE.DEVIATION.value)})=ABS({XLS.CELL(ROW, HEADERS_NOVISIBLE.SPECIFICATION.value)}),"4 UGB1",
    IF(AND(ABS({XLS.CELL(ROW, HEADERS_NOVISIBLE.DEVIATION.value)})>ABS({XLS.CELL(ROW, HEADERS_NOVISIBLE.SPECIFICATION.value)}),ABS({XLS.CELL(ROW, HEADERS_NOVISIBLE.DEVIATION.value)})-ABS({XLS.CELL(ROW, HEADERS_NOVISIBLE.UNCERTAINTY.value)})<ABS({XLS.CELL(ROW, HEADERS_NOVISIBLE.SPECIFICATION.value)})),"5 UGB2",
    IF(AND(ABS({XLS.CELL(ROW, HEADERS_NOVISIBLE.DEVIATION.value)})>ABS({XLS.CELL(ROW, HEADERS_NOVISIBLE.SPECIFICATION.value)}),ABS({XLS.CELL(ROW, HEADERS_NOVISIBLE.DEVIATION.value)})-ABS({XLS.CELL(ROW, HEADERS_NOVISIBLE.UNCERTAINTY.value)})=ABS({XLS.CELL(ROW, HEADERS_NOVISIBLE.SPECIFICATION.value)})),"6 UGB2",
    IF(ABS({XLS.CELL(ROW, HEADERS_NOVISIBLE.DEVIATION.value)})-ABS({XLS.CELL(ROW, HEADERS_NOVISIBLE.UNCERTAINTY.value)})>ABS({XLS.CELL(ROW, HEADERS_NOVISIBLE.SPECIFICATION.value)}),"7 Fail",
))))))))))'''
        return FORMULA

    def EXCEL_FORMULA_RESULT_HEADER() -> str:
        FORMULA: str = \
f'''=
IF(COUNTIF({XLS.get_column_letter(HEADERS_NOVISIBLE.RESULT.value)}:{XLS.get_column_letter(HEADERS_NOVISIBLE.RESULT.value)},"n.m.")>0,"n.m.",
IF(COUNTIF({XLS.get_column_letter(HEADERS_NOVISIBLE.RESULT.value)}:{XLS.get_column_letter(HEADERS_NOVISIBLE.RESULT.value)},"7 Fail")>0,"FAIL",
IF(COUNTIF({XLS.get_column_letter(HEADERS_NOVISIBLE.RESULT.value)}:{XLS.get_column_letter(HEADERS_NOVISIBLE.RESULT.value)},"*UGB2")>0,"UGB2",
IF(COUNTIF({XLS.get_column_letter(HEADERS_NOVISIBLE.RESULT.value)}:{XLS.get_column_letter(HEADERS_NOVISIBLE.RESULT.value)},"*UGB1")>0,"UGB1",
"PASS"
))))'''
        return FORMULA

    def EXCEL_FORMULA_INDICATION(ROW: int) -> str:
        FORMULA: str = \
f'''=
IF({XLS.CELL(ROW, HEADERS_NOVISIBLE.RESULT.value)}="n.i.","n.i.",
IF({XLS.CELL(ROW, HEADERS_NOVISIBLE.RESULT.value)}="n.m.","n.m.",
IF({XLS.CELL(ROW, HEADERS_NOVISIBLE.RESULT.value)}="3 UGB1","u",
IF({XLS.CELL(ROW, HEADERS_NOVISIBLE.RESULT.value)}="4 UGB1","u",
IF({XLS.CELL(ROW, HEADERS_NOVISIBLE.RESULT.value)}="5 UGB2","ü",
IF({XLS.CELL(ROW, HEADERS_NOVISIBLE.RESULT.value)}="6 UGB2","ü",
IF({XLS.CELL(ROW, HEADERS_NOVISIBLE.RESULT.value)}="7 Fail","FAIL",
"")))))))'''
        return FORMULA

    def EXCEL_FORMULA_TYPA(ROW: int, COL_INI: int, COL_FIN: int) -> str:
        '''
        =IF(COUNT(AL62:AV62)<2;0;STDEV.S(AL62:AV62)/SQRT((COUNT(AL62:AV62))))
        '''
        return f'''=IF(COUNT({XLS.CELL(ROW, COL_INI)}:{XLS.CELL(ROW, COL_FIN)})<2,0,_xlfn.STDEV.S({XLS.CELL(ROW, COL_INI)}:{XLS.CELL(ROW, COL_FIN)})/SQRT((COUNT({XLS.CELL(ROW, COL_INI)}:{XLS.CELL(ROW, COL_FIN)}))))'''

    def EXCEL_FORMULA_DIGITS(ROW: int) -> str:
        FORMULA: str = \
f'''=
IF({XLS.CELL(ROW, HEADERS_NOVISIBLE.RESULT.value)}="n.i.","",
IF({XLS.CELL(ROW, HEADERS_NOVISIBLE.RESULT.value)}="n.m.","",
IF(INT(RIGHT(TEXT({XLS.CELL(ROW, HEADERS_NOVISIBLE.UNCERTAINTY.value)},"0E+000"),4))>0,
TEXT(ROUND({XLS.CELL(ROW, HEADERS_NOVISIBLE.DEVIATION.value)},-INT(RIGHT(TEXT({XLS.CELL(ROW, HEADERS_NOVISIBLE.UNCERTAINTY.value)},"0E+000"),4))+1)/10^INT(RIGHT(TEXT({XLS.CELL(ROW, HEADERS_NOVISIBLE.UNCERTAINTY.value)},"0E+000"),4)),"0,0") & "E+" & INT(RIGHT(TEXT({XLS.CELL(ROW, HEADERS_NOVISIBLE.UNCERTAINTY.value)},"0E+000"),4)),
TEXT(ROUND({XLS.CELL(ROW, HEADERS_NOVISIBLE.DEVIATION.value)},-INT(RIGHT(TEXT({XLS.CELL(ROW, HEADERS_NOVISIBLE.UNCERTAINTY.value)},"0E+000"),4))+1),"0," & REPT(0,-INT(RIGHT(TEXT({XLS.CELL(ROW, HEADERS_NOVISIBLE.UNCERTAINTY.value)},"0E+000"),4))+1))
) & CHAR(32) & {XLS.CELL(ROW, HEADERS_NOVISIBLE.DEVIATION.value+1)}
))
'''
        return FORMULA

    def EXCEL_FORMULA_FREEDOMDEGREES() -> float:
        '''
        FREEDOM_DEGREES
        =IF(
        (IF(BN63<>0;(BN63*1)^4/COUNT(I63:R63)-1;0)+IF(BP63<>0;(BP63*1)^4/COUNT(T63:AC63)-1;0))=0;0;
        BF63^4/(IF(BN63<>0;(BN63*1)^4/COUNT(I63:R63)-1;0)+IF(BP63<>0;(BP63*1)^4/COUNT(T63:AC63)-1;0))
        )
        '''

    def FORMAT(REPORT: XLS.XLSREPORT) -> None:
        '''
        '''
        if REPORT.WS.page_setup.fitToPage == False:
            REPORT.WS.page_setup.fitToPage = True
        REPORT.WS.page_setup.fitToPage = 1
        REPORT.WS.page_setup.fitToHeight = False
        REPORT.WS.print_area = 'A1:G1048576'
        footer = "PAGE &P OF &N"
        # footer = header_footer.HeaderFooterItem()
        # footer.right()
        # REPORT.WS.HeaderFooter.oddFooter.left.font = Font(name='Arial', size=12, bold=True)
        REPORT.WS.HeaderFooter.oddFooter.left.text = footer
        for col in range(1,7):
            REPORT.COL_WIDTH(col, 20)
        REPORT.COL_WIDTH(6, 5) # RESULT INDICATION
        REPORT.COL_WIDTH(7, 5) # ACREDITED
        REPORT.COL_WIDTH(8, 15) # PROCEDURE
        # REPORT.SAVE()
        # Crear una regla de formato condicional usando una fórmula
        # Suponiendo que quieres aplicarlo a la columna H (la columna 8)
        # rule = FormulaRule(formula=[f'OR($K1="7 Fail", $K1="5 UGB2", $K1="6 UGB2"'], fill=red_fill, stopIfTrue=True)

        # Aplicar el formato condicional al rango de celdas
        # Aquí se especifica un rango amplio, pero puedes ajustar '1048576' al número de tu última fila
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        # print(f'OR(${XLS.CELL(1, HEADERS_NOVISIBLE.RESULT.value)}="FAIL"),${XLS.CELL(1, HEADERS_NOVISIBLE.RESULT.value)}="5 UGB2",${XLS.CELL(1, HEADERS_NOVISIBLE.RESULT.value)}="6 UGB2"))')
        rule = FormulaRule(formula=[f'OR(${XLS.CELL(1, HEADERS_NOVISIBLE.RESULT.value)}="7 Fail",${XLS.CELL(1, HEADERS_NOVISIBLE.RESULT.value)}="5 UGB2",${XLS.CELL(1, HEADERS_NOVISIBLE.RESULT.value)}="6 UGB2")'], fill=red_fill, stopIfTrue=True)
        REPORT.WS.conditional_formatting.add('H1:H1048576', rule)

    def HEADER(REPORT: XLS.XLSREPORT, MODEL_ID):
        SQL = SQL_BY_ROW("MODELS", "Id", MODEL_ID)[0]
        ##
        img_results_path = os.path.join("resources",r"R&S Logo - Complete.png")
        img_results = XLS.Image(img_results_path)
        img_results.width = 263
        img_results.height = 68
        REPORT.WS.add_image(img_results, XLS.CELL(1, 4))
        ##
        REPORT.WR_HEADER(REPORT.ROW, 1, "MANUFACTURER:")
        REPORT.WR_HEADER(REPORT.ROW, 8, "RESULT:")
        REPORT.WR(REPORT.ROW, 2, SQL['MANUFACTURER'])
        REPORT.ROW_INC()
        REPORT.WR_HEADER(REPORT.ROW, 1, "MODEL:")
        REPORT.WR(REPORT.ROW, 2, SQL['MODEL'])
        REPORT.WR(REPORT.ROW, 8, EXCEL_FORMULA_RESULT_HEADER())
        REPORT.ROW_INC()
        REPORT.WR_HEADER(REPORT.ROW, 1, "SERIAL Id:")
        # REPORT.WR(REPORT.ROW, 2, SQL['SERIAL'])
        REPORT.ROW_INC()
        REPORT.WR_HEADER(REPORT.ROW, 1, "DATE:")
        REPORT.WR(REPORT.ROW, 2, datetime.now().strftime(r'%Y-%m-%d'))
        REPORT.ROW_INC()
        REPORT.WR_HEADER(REPORT.ROW, 1, "CALIBRATION Id:")
        REPORT.ROW_INC()
        REPORT.LOW_BORDER(REPORT.ROW, col_fin=8)
        REPORT.ROW_INC(2)
        REPORT.SHEET_HEAD(6)
        REPORT.SAVE()

    def PAGE3(REPORT: XLS.XLSREPORT) -> None:
        '''
        '''
        FONT1 = XLS.Font(name='Calibri', size=12, bold=False)
        FONT2 = XLS.Font(name='Calibri', size=9, bold=False)
        # 
        REPORT.WR_TITLE(REPORT.ROW, 1, r"Comments on the measured results")
        REPORT.ROW_INC()
        REPORT.WR(REPORT.ROW, 1, r"The measurement results in the test report stated below have been tested for compliance with the given specifications and marked if necessary. In doing so, the associated uncertainty of measurement has been taken into acount.")
        REPORT.WS.merge_cells(
            start_row=REPORT.ROW, 
            end_row=REPORT.ROW+1,
            start_column=1,
            end_column=5
            )
        # BUG: Create Function in pydeveloptools
        REPORT.WS.cell(REPORT.ROW, 1).alignment = XLS.Alignment(wrap_text=True, vertical="top")
        REPORT.ROW_INC(2)
        img_results_path = os.path.join("resources", r"Results.jpg")
        img_results = XLS.Image(img_results_path)
        img_results.width = 491
        img_results.height = 227
        REPORT.WS.add_image(img_results, XLS.CELL(REPORT.ROW, 1))
        # REPORT.IMAGE_INSERT(REPORT.ROW, 1, )
        REPORT.ROW_INC(13)
        REPORT.WR_TITLE(REPORT.ROW, 1, r"The following abbrevations may be used in this certificate:")
        REPORT.ROW_INC()
        REPORT.WR(REPORT.ROW, 1, r"¹", FONT1)
        REPORT.WR(REPORT.ROW, 2, r"Measurement results that are not covered by the DAkkS accreditation.", FONT2)
        REPORT.ROW_INC()
        REPORT.ROW_WIDTH(REPORT.ROW, 5)
        REPORT.ROW_INC()
        REPORT.WR(REPORT.ROW, 1, "{a}", FONT1)
        REPORT.WR(REPORT.ROW, 2, r'No measurement uncertainty stated because the errors always add together.', FONT2)
        REPORT.ROW_INC()
        REPORT.WR(REPORT.ROW, 2, r'So it is sure that a measurement result evaluted as "PASS" is pass.', FONT2)
        REPORT.ROW_INC()
        REPORT.ROW_WIDTH(REPORT.ROW, 5)
        REPORT.ROW_INC()
        REPORT.WR(REPORT.ROW, 1, "{b}", FONT1)
        REPORT.WR(REPORT.ROW, 2, r"The measurement uncertainty depends on the measurement result. ", FONT2)
        REPORT.ROW_INC()
        REPORT.WR(REPORT.ROW, 2, r"The stated measurement uncertainty is valid for the close area around the specification. ", FONT2)
        REPORT.ROW_INC()
        REPORT.WR(REPORT.ROW, 2, r"Measurement results outside the close area have a higher measurement uncertainty but are within the specification.", FONT2)
        REPORT.ROW_INC()
        REPORT.ROW_WIDTH(REPORT.ROW, 5)
        REPORT.ROW_INC()
        REPORT.WR(REPORT.ROW, 1, r"{c} ,  ²", FONT1)
        REPORT.WR(REPORT.ROW, 2, r"Functional test, therefore no measurement uncertainty is stated.", FONT2)
        REPORT.ROW_INC()
        REPORT.ROW_WIDTH(REPORT.ROW, 5)
        REPORT.ROW_INC()
        REPORT.WR(REPORT.ROW, 1, r"{d}", FONT1)
        REPORT.WR(REPORT.ROW, 2, r"Typical value, refer to performance test.", FONT2)
        REPORT.ROW_INC()
        REPORT.ROW_WIDTH(REPORT.ROW, 5)
        REPORT.ROW_INC()
        REPORT.WR(REPORT.ROW, 1, r"{e}", FONT1)
        REPORT.WR(REPORT.ROW, 2, r"The measurement uncertainty is taken into account when setting the measuring system.", FONT2)
        REPORT.ROW_INC()
        REPORT.ROW_WIDTH(REPORT.ROW, 5)
        REPORT.ROW_INC()
        REPORT.WR(REPORT.ROW, 1, r"{g}", FONT1)
        REPORT.WR(REPORT.ROW, 2, r"Verification of specified requirements, non-accredited measurements. ", FONT2)
        REPORT.ROW_INC()
        REPORT.WR(REPORT.ROW, 2, r"Technical operation that consist of the determination of one or more characteristics to", FONT2)
        REPORT.ROW_INC()
        REPORT.WR(REPORT.ROW, 2, r"a specified procedure (formerly {f}).", FONT2)
        REPORT.ROW_INC()
        REPORT.ROW_WIDTH(REPORT.ROW, 5)
        REPORT.ROW_INC()
        REPORT.WR(REPORT.ROW, 1, "DL , DT", FONT1)
        REPORT.WR(REPORT.ROW, 2, r"Data Limit for symmetrical tolerance limits", FONT2)
        REPORT.ROW_INC()
        REPORT.ROW_WIDTH(REPORT.ROW, 5)
        REPORT.ROW_INC()
        REPORT.WR(REPORT.ROW, 1, "UGB", FONT1)
        REPORT.WR(REPORT.ROW, 2, r"Uncertainty guard band: Measuring uncertainty violates the data sheet tolerance", FONT2)
        REPORT.ROW_INC()
        REPORT.ROW_WIDTH(REPORT.ROW, 5)
        REPORT.ROW_INC()
        REPORT.WR(REPORT.ROW, 1, "UGB1 , u", FONT1)
        REPORT.WR(REPORT.ROW, 2, r"Measurement results marked as UGB1 show conformity with a probability of >50 % and <95 %.", FONT2)
        REPORT.ROW_INC()
        REPORT.WR(REPORT.ROW, 1, "UGB2 , ü", FONT1)
        REPORT.WR(REPORT.ROW, 2, r"Measurement results marked as UGB2 show non-conformity with a probability of >50 % and <95 %.", FONT2)
        REPORT.ROW_INC()
        REPORT.ROW_WIDTH(REPORT.ROW, 5)
        REPORT.ROW_INC()
        REPORT.WR(REPORT.ROW, 1, "FAIL , f", FONT1)
        REPORT.WR(REPORT.ROW, 2, r"Measurement results marked as FAIL show non-conformity", FONT2)
        REPORT.ROW_INC()
        REPORT.ROW_WIDTH(REPORT.ROW, 5)
        REPORT.ROW_INC()
        REPORT.WR(REPORT.ROW, 1, "n. i.", FONT1)
        REPORT.WR(REPORT.ROW, 2, r"not installed: Does not apply due to instrument configuration", FONT2)
        REPORT.ROW_INC()
        REPORT.ROW_WIDTH(REPORT.ROW, 5)
        REPORT.ROW_INC()
        REPORT.WR(REPORT.ROW, 1, "n. m.", FONT1)
        REPORT.WR(REPORT.ROW, 2, r"not measured", FONT2)
        REPORT.ROW_INC()
        REPORT.ROW_WIDTH(REPORT.ROW, 5)
        REPORT.ROW_INC()
        REPORT.WR(REPORT.ROW, 1, "ref.", FONT1)
        REPORT.WR(REPORT.ROW, 2, r"Reference value, used for relative measurements", FONT2)
        REPORT.ROW_INC(2)
        REPORT.PAGE_BREAK(REPORT.ROW)
        REPORT.ROW_INC()
        REPORT.SAVE()

    def SUMMARY(REPORT: XLS.XLSREPORT) -> None:
        REPORT.SHEET_NEW("Summary")
        REPORT.ROW = 2
        REPORT.COL_WIDTH(8, 30)
        # REPORT.SHEET_SELECT('Summary')

        ##
        REPORT.WR_HEADER(REPORT.ROW, 1, "Certificate No.:"); 
        REPORT.SET_RANGE_NAME(REPORT.ROW, 2, 'CERTIFICATE')
        REPORT.ROW_INC(1)
        REPORT.WR_HEADER(REPORT.ROW, 1, "Date:"); REPORT.ROW_INC(2)

        REPORT.WR_HEADER(REPORT.ROW, 1, "Type Device:"); REPORT.ROW_INC(1)
        REPORT.WR_HEADER(REPORT.ROW, 1, "Manufacturer:"); REPORT.ROW_INC(1)
        REPORT.WR_HEADER(REPORT.ROW, 1, "Model:"); REPORT.ROW_INC(1)
        REPORT.WR_HEADER(REPORT.ROW, 1, "Material No.:"); REPORT.ROW_INC(1)
        REPORT.WR_HEADER(REPORT.ROW, 1, "Serial Id:"); REPORT.ROW_INC(1)
        REPORT.WR_HEADER(REPORT.ROW, 1, "Inventory Id"); REPORT.ROW_INC(2)

        REPORT.WR_HEADER(REPORT.ROW, 1, "Scope of Cal."); REPORT.ROW_INC(1)
        REPORT.WR_HEADER(REPORT.ROW, 1, "Place of Cal."); REPORT.ROW_INC(1)
        REPORT.WR_HEADER(REPORT.ROW, 1, "Technician Id"); REPORT.ROW_INC(2)

        REPORT.WR_HEADER(REPORT.ROW, 1, "Temperature [ºC]"); REPORT.ROW_INC(1)
        REPORT.WR_HEADER(REPORT.ROW, 1, "Humidity [%HR]"); REPORT.ROW_INC(1)
        REPORT.WR_HEADER(REPORT.ROW, 1, "Result:"); REPORT.ROW_INC(2)

        REPORT.WR_HEADER(REPORT.ROW, 1, "Customer:"); REPORT.ROW_INC(1)
        REPORT.WR_HEADER(REPORT.ROW, 1, "Customer Address:")
        REPORT.LOW_BORDER(REPORT.ROW,1,10)
        REPORT.ROW_INC(6)

        REPORT.WR_HEADER(REPORT.ROW, 1, "Standards Used:")
        REPORT.LOW_BORDER(REPORT.ROW,1,10)
        REPORT.ROW_INC(1)
        # REPORT.SAVE()

    def TYPB_TBL(REPORT: XLS.XLSREPORT) -> None:
        REPORT.SHEET_NEW("TYP_B")
        REPORT.ROW = 1

    st.text("")
    col12, col22 = st.columns(2)

    with col12:

        holder = st.empty()
        if holder.button("GET TEMPLATE .xlsx", use_container_width=True):
            # with st.spinner('Creating .xlsx file...'):
            
            MODEL_ID = CURRENT_TEMPLATE["MODEL_ID"]

            if not os.path.exists("REPORTS"):
                os.mkdir('REPORTS')
            path_file = os.path.join("REPORTS",f"{MODEL_ID}.xlsx")
            if os.path.exists(path_file):
                os.remove(path_file)
            
            ## REPORT
            REPORT = XLS.XLSREPORT(path_file, worksheet_name='Test-Report')
            FORMAT(REPORT)
            HEADER(REPORT, MODEL_ID)
            PAGE3(REPORT)

            MODEL_DB = SQL_SELECT_DB("MODELS", MODEL_ID)
            SPECIFICATION_DICT = json.loads(MODEL_DB)['SPECIFICATIONS']
            STANDARDS = []

            ## TEST
            for test in CURRENT_DB["TEST_LIST"]:
                current_test = TEMPLATE.TEST(**test)
                PROCEDURE_ID = current_test.PROCEDURE_ID
                SQL_PROCEDURE = SQL_BY_ROW("PROCEDURES", "Id", PROCEDURE_ID)[0]
                if isinstance(SQL_PROCEDURE['DB'], dict): PROCEDURE_DB = SQL_PROCEDURE['DB']
                if isinstance(SQL_PROCEDURE['DB'], str): PROCEDURE_DB = json.loads(SQL_PROCEDURE['DB'])
                PROCEDURE_PYDATA = SQL_PROCEDURE['PYDATA']

                ## STANDARDS
                for standard in PROCEDURE_DB["STANDARDS"]:
                    if standard not in STANDARDS:
                        STANDARDS.append(standard)

                ## PYDATA
                MODULE = {}
                try:
                    exec(PROCEDURE_PYDATA, MODULE)
                except Exception as e:
                    print("ERROR / ImportLib:", PROCEDURE_ID)
                    print(e)
                # CMC_DICT = json.loads(PROCEDURE_DB)['CMC']
                CMC_DF = pd.DataFrame(PROCEDURE_DB['CMC'])
                if PROCEDURE_ID  in SPECIFICATION_DICT:
                    SPECIFICATION_DF = pd.DataFrame(SPECIFICATION_DICT[PROCEDURE_ID]) #, columns=['RANGE1_MIN', 'RANGE1_MAX','RANGE2_MIN','RANGE2_MAX','RESOLUTION','C1','C2','C3','EVALUATION',])
                else:
                    SPECIFICATION_DF = None
                    st.toast(f"SPECIFICATION <{PROCEDURE_ID}> IS EMPTY")
                # print(SPECIFICATION_DF)

                ## TEST TITLE
                TEST_TITLE = current_test.TEST
                REPORT.WR_TITLE(REPORT.ROW, 1, f"{CURRENT_DB["TEST_LIST"].index(test)+1} - {TEST_TITLE}")
                REPORT.ROW_INC()

                ## PARAMETERS
                PARAMETERS = current_test.PARAMETERS
                # print(PARAMETERS)
                if PARAMETERS: # Hay casos que puede estar vacio
                    REPORT.WR(REPORT.ROW, 1, PARAMETERS, XLS.FONTS.CAPTION.value)
                    # REPORT.ROW_WIDTH(REPORT.ROW, 10)
                    REPORT.ROW_INC()

                ## CALIBRATION CONTENT
                REPORT.WR_HEADERS(REPORT.ROW, 1, ["TEST PARAMETERS", "", "RESULT MEAS.", "LIMIT OF ERROR", "UNCERTAINTY"])
                for header in HEADERS_NOVISIBLE:
                    REPORT.WR_HEADER(REPORT.ROW, header.value, header.name)
                REPORT.LOW_BORDER(REPORT.ROW, HEADERS_NOVISIBLE.ACQ_MEASURE.value,HEADERS_NOVISIBLE.ACQ_MEASURE.value+10)
                REPORT.LOW_BORDER(REPORT.ROW, HEADERS_NOVISIBLE.ACQ_INDICATION.value,HEADERS_NOVISIBLE.ACQ_INDICATION.value+10)
                if MODULE.get("RESULT_ID"):
                    REPORT.WR_HEADER(REPORT.ROW, 3, MODULE["RESULT_ID"])
                REPORT.ROW_INC()

                CALIBRATION_DF = pd.DataFrame(current_test.CALIBRATION)
                for loc in CALIBRATION_DF.index:
                    ROW_DATA = dict(CALIBRATION_DF.loc[loc])
                    ## SPECIFICATION
                    if isinstance(SPECIFICATION_DF, pd.DataFrame):
                        # RESOLUTION = TABLE_DATA.FILTER_VALUE(SPECIFICATION_DF, 'RESOLUTION', ROW_DATA['VALUE1'], ROW_DATA['VALUE2'])
                        RESOLUTION = SPECIFICATION_DF[SPECIFICATION_DF['RANGE1_MAX']==ROW_DATA['RANGE']]['RESOLUTION'].max()
                    # if MODULE.get("SPECIFICATION"):
                    #     SPECIFICATION = MODULE['SPECIFICATION'](SPECIFICATION_DF, ROW_DATA['VALUE1'], ROW_DATA['VALUE2'])
                    # else:
                        SPECIFICATION = TABLE_DATA.GET_VALUE(SPECIFICATION_DF, ROW_DATA['VALUE1'], ROW_DATA['VALUE2'])
                    else:
                        RESOLUTION = None
                        SPECIFICATION = None
                    ## CMC
                    # if MODULE.get("CMC"):
                    #     CMC = MODULE['CMC'](CMC_DF, ROW_DATA['VALUE1'], ROW_DATA['VALUE2'])
                    # else:
                    if PROCEDURE_DB['REPORT_FORMAT'].get('ABSOLUTE_VALUES'): ABS = PROCEDURE_DB['REPORT_FORMAT']['ABSOLUTE_VALUES']
                    else: ABS = False
                    CMC = TABLE_DATA.GET_VALUE(CMC_DF, ROW_DATA['VALUE1'], ROW_DATA['VALUE2'], ABS=ABS)

                    ## REPORT
                    REPORT.WR(REPORT.ROW, 3, EXCEL_FORMULA_DIGITS(REPORT.ROW))
                    REPORT.WR(REPORT.ROW, 4, f'={XLS.get_column_letter(HEADERS_NOVISIBLE.SPECIFICATION.value)}{REPORT.ROW}')
                    REPORT.WS.cell(REPORT.ROW, 4).number_format = f'0.0E+0 "{PROCEDURE_DB['REPORT_FORMAT']['UNITS'][HEADERS_NOVISIBLE.SPECIFICATION.name]}"'
                    REPORT.WR(REPORT.ROW, 5, f'={XLS.get_column_letter(HEADERS_NOVISIBLE.UNCERTAINTY.value)}{REPORT.ROW}')
                    REPORT.WS.cell(REPORT.ROW, 5).number_format = f'0.0E+0 "{PROCEDURE_DB['REPORT_FORMAT']['UNITS'][HEADERS_NOVISIBLE.UNCERTAINTY.name]}"'
                    REPORT.WR(REPORT.ROW, 6, EXCEL_FORMULA_INDICATION(REPORT.ROW))
                    formula = rf'=IF(OR({XLS.get_column_letter(HEADERS_NOVISIBLE.CMC.value)}{REPORT.ROW}=0,ISBLANK({XLS.get_column_letter(HEADERS_NOVISIBLE.CMC.value)}{REPORT.ROW})),"{chr(123)}g{chr(125)}","")'
                    REPORT.WR(REPORT.ROW, 7, formula)
                    try:
                        REPORT.WR(REPORT.ROW, 1, PROCEDURE_DB['REPORT_FORMAT']['PARAMETERS']['PARAMETERS'].format(**ROW_DATA), FONT=XLS.FONTS.CAPTION.value)
                    except:
                        REPORT.WR(REPORT.ROW, 1, "RANGE: {RANGE} | NOMINAL: {VALUE1}".format(**ROW_DATA), FONT=XLS.FONTS.CAPTION.value)
                    
                    ## NO VISIBLE
                    REPORT.WR(REPORT.ROW, HEADERS_NOVISIBLE.PROCEDURE_ID.value, PROCEDURE_ID)
                    REPORT.WR(REPORT.ROW, HEADERS_NOVISIBLE.MEAS_TYPE.value, PROCEDURE_DB['REPORT_FORMAT']['RESULT_TYPE'])
                    REPORT.WR_SCI_NUMBER(REPORT.ROW, HEADERS_NOVISIBLE.RESULT.value, EXCEL_FORMULA_RESULT(REPORT.ROW))
                    REPORT.WR_SCI_NUMBER(REPORT.ROW, HEADERS_NOVISIBLE.RANGE.value, ROW_DATA['RANGE'])
                    REPORT.WR_SCI_NUMBER(REPORT.ROW, HEADERS_NOVISIBLE.VALUE1.value, ROW_DATA['VALUE1'])
                    REPORT.WR_SCI_NUMBER(REPORT.ROW, HEADERS_NOVISIBLE.VALUE2.value, ROW_DATA['VALUE2'])
                    REPORT.WR_SCI_NUMBER(REPORT.ROW, HEADERS_NOVISIBLE.MEASURE.value, f'=IF(COUNT({XLS.get_column_letter(HEADERS_NOVISIBLE.ACQ_MEASURE.value)}{REPORT.ROW}:{XLS.get_column_letter(HEADERS_NOVISIBLE.ACQ_MEASURE.value+9)}{REPORT.ROW})>0,AVERAGE({XLS.get_column_letter(HEADERS_NOVISIBLE.ACQ_MEASURE.value)}{REPORT.ROW}:{XLS.get_column_letter(HEADERS_NOVISIBLE.ACQ_MEASURE.value+9)}{REPORT.ROW}),FALSE)')
                    REPORT.WR_SCI_NUMBER(REPORT.ROW, HEADERS_NOVISIBLE.INDICATION.value, f'=IF(COUNT({XLS.get_column_letter(HEADERS_NOVISIBLE.ACQ_INDICATION.value)}{REPORT.ROW}:{XLS.get_column_letter(HEADERS_NOVISIBLE.ACQ_INDICATION.value+9)}{REPORT.ROW})>0,AVERAGE({XLS.get_column_letter(HEADERS_NOVISIBLE.ACQ_INDICATION.value)}{REPORT.ROW}:{XLS.get_column_letter(HEADERS_NOVISIBLE.ACQ_INDICATION.value+9)}{REPORT.ROW}),{XLS.CELL(REPORT.ROW, HEADERS_NOVISIBLE.VALUE1.value)})')
                    REPORT.WR_SCI_NUMBER(REPORT.ROW, HEADERS_NOVISIBLE.DEVIATION.value, f'=IF({XLS.CELL(REPORT.ROW, HEADERS_NOVISIBLE.MEASURE.value)}=FALSE,FALSE,{XLS.CELL(REPORT.ROW, HEADERS_NOVISIBLE.MEASURE.value)}-{XLS.CELL(REPORT.ROW, HEADERS_NOVISIBLE.INDICATION.value)})')
                    REPORT.WR_SCI_NUMBER(REPORT.ROW, HEADERS_NOVISIBLE.SPECIFICATION.value, SPECIFICATION)
                    REPORT.WR_SCI_NUMBER(REPORT.ROW, HEADERS_NOVISIBLE.CMC.value, CMC)
                    REPORT.WR_SCI_NUMBER(REPORT.ROW, HEADERS_NOVISIBLE.RESOLUTION.value, RESOLUTION)
                    REPORT.WR_SCI_NUMBER(REPORT.ROW, HEADERS_NOVISIBLE.UNCERTAINTY.value, f'=IF(ISBLANK({XLS.get_column_letter(HEADERS_NOVISIBLE.CMC.value)}{REPORT.ROW}),{XLS.get_column_letter(HEADERS_NOVISIBLE.U.value)}{REPORT.ROW},MAX({XLS.get_column_letter(HEADERS_NOVISIBLE.CMC.value)}{REPORT.ROW},{XLS.get_column_letter(HEADERS_NOVISIBLE.U.value)}{REPORT.ROW}))')
                    REPORT.WR_SCI_NUMBER(REPORT.ROW, HEADERS_NOVISIBLE.U.value, f'=2*SQRT(SUM(({XLS.CELL(REPORT.ROW, HEADERS_NOVISIBLE.RESOLUTION.value)}/(2*SQRT(3)))^2,{XLS.CELL(REPORT.ROW, HEADERS_NOVISIBLE.TYPB.value)}^2,{XLS.CELL(REPORT.ROW, HEADERS_NOVISIBLE.TYPA_MEAS.value)}^2,{XLS.CELL(REPORT.ROW, HEADERS_NOVISIBLE.TYPA_INDI.value)}^2))')
                    # REPORT.WR_SCI_NUMBER(REPORT.ROW, HEADERS_NOVISIBLE.TYPA.value, f'=UNCER_TIP_A({XLS.get_column_letter(HEADERS_NOVISIBLE.TYPA.value+2)}{REPORT.ROW}:{XLS.get_column_letter(HEADERS_NOVISIBLE.TYPA.value+7)}{REPORT.ROW})')
                    REPORT.WR_SCI_NUMBER(REPORT.ROW, HEADERS_NOVISIBLE.TYPA_MEAS.value, EXCEL_FORMULA_TYPA(REPORT.ROW, HEADERS_NOVISIBLE.ACQ_MEASURE.value, HEADERS_NOVISIBLE.ACQ_MEASURE.value+9))   
                    REPORT.WR_SCI_NUMBER(REPORT.ROW, HEADERS_NOVISIBLE.TYPA_INDI.value, EXCEL_FORMULA_TYPA(REPORT.ROW, HEADERS_NOVISIBLE.ACQ_INDICATION.value, HEADERS_NOVISIBLE.ACQ_INDICATION.value+9))   
                    
                    ## UNITS
                    try:
                        REPORT.WR(REPORT.ROW, HEADERS_NOVISIBLE.RANGE.value+1, PROCEDURE_DB['REPORT_FORMAT']['UNITS'][HEADERS_NOVISIBLE.RANGE.name])
                        REPORT.WR(REPORT.ROW, HEADERS_NOVISIBLE.VALUE1.value+1, PROCEDURE_DB['REPORT_FORMAT']['UNITS'][HEADERS_NOVISIBLE.VALUE1.name])
                        REPORT.WR(REPORT.ROW, HEADERS_NOVISIBLE.VALUE2.value+1, PROCEDURE_DB['REPORT_FORMAT']['UNITS'][HEADERS_NOVISIBLE.VALUE2.name])
                        REPORT.WR(REPORT.ROW, HEADERS_NOVISIBLE.INDICATION.value+1, PROCEDURE_DB['REPORT_FORMAT']['UNITS'][HEADERS_NOVISIBLE.MEASURE.name])
                        REPORT.WR(REPORT.ROW, HEADERS_NOVISIBLE.MEASURE.value+1, PROCEDURE_DB['REPORT_FORMAT']['UNITS'][HEADERS_NOVISIBLE.MEASURE.name])
                        REPORT.WR(REPORT.ROW, HEADERS_NOVISIBLE.DEVIATION.value+1, PROCEDURE_DB['REPORT_FORMAT']['UNITS'][HEADERS_NOVISIBLE.MEASURE.name])
                        REPORT.WR(REPORT.ROW, HEADERS_NOVISIBLE.SPECIFICATION.value+1, PROCEDURE_DB['REPORT_FORMAT']['UNITS'][HEADERS_NOVISIBLE.SPECIFICATION.name])
                        REPORT.WR(REPORT.ROW, HEADERS_NOVISIBLE.CMC.value+1, PROCEDURE_DB['REPORT_FORMAT']['UNITS'][HEADERS_NOVISIBLE.CMC.name])
                        REPORT.WR(REPORT.ROW, HEADERS_NOVISIBLE.UNCERTAINTY.value+1, PROCEDURE_DB['REPORT_FORMAT']['UNITS'][HEADERS_NOVISIBLE.UNCERTAINTY.name])
                    except:
                        print("ERROR UNITS:", PROCEDURE_ID)
                    
                    ## PROTECTION
                    for col in range(HEADERS_NOVISIBLE.PROCEDURE_ID.value, HEADERS_NOVISIBLE.TYPA_INDI.value):
                        REPORT.WS.cell(REPORT.ROW, col).protection = Protection(locked=True)
                    for col in range(HEADERS_NOVISIBLE.ACQ_MEASURE.value, HEADERS_NOVISIBLE.ACQ_MEASURE.value+9):
                        REPORT.WS.cell(REPORT.ROW, col).protection = Protection(locked=False)
                    
                    REPORT.ROW_INC()
                    
                del MODULE  
                REPORT.ROW_INC()
                REPORT.PAGE_BREAK(REPORT.ROW)
                # REPORT.SAVE()

            REPORT.WS.protection.sheet = True
            REPORT.SAVE()
            SUMMARY(REPORT)
            
            for standard in STANDARDS:
                REPORT.WR(REPORT.ROW, 1, standard)
                REPORT.ROW_INC()

            ## standards
            REPORT.SAVE()
            
            holder.download_button(
                label="📩 DOWNLOAD .xlsx File",
                data=open(path_file, "rb").read(),
                file_name=f'{MODEL_ID}.xlsx',
                mime="calcs/xlsx",
                use_container_width=True
            )
            os.remove(path_file)


## PAGE
## __________________________________________________________________________________________________

SIDEBAR()

st.text('TEMPLATE Id')

SQL = SQL_TEMPLATES_VIEW(st.session_state.TEMPLATES)
DATAFRAME = pd.DataFrame(SQL)

FLTR_DEVICE: str = None
FLTR_MANUFACTURER: str = None
FLTR_ID: str = None

def get_filter() -> pd.DataFrame:
    df_filtered = DATAFRAME.copy()
    # for filter in [FLTR_DEVICE, FLTR_MANUFACTURER, FLTR_ID]:
    #     if filter == None or 
    if FLTR_DEVICE:
        df_filtered = df_filtered[df_filtered['DEVICE_TYPE']==FLTR_DEVICE]
    if FLTR_MANUFACTURER:
        df_filtered = df_filtered[df_filtered['MANUFACTURER']==FLTR_MANUFACTURER]
    if FLTR_ID:
        df_filtered = df_filtered[df_filtered['Id']==FLTR_ID]
    return df_filtered

col12, col22 = st.columns(2)

with col12:
    holder_template = st.empty()
    TEMPLATE_ID = holder_template.selectbox("TEMPLATE Id", options=DATAFRAME['Id'].to_list(), index=None, label_visibility='collapsed')

with col22:
    with st.popover(USUAL_ICONS.EXPANDER.value):
        FLTR_DEVICE = st.selectbox("DEVICE TYPE", options=get_filter()['DEVICE_TYPE'].unique().tolist(), index=None)
        FLTR_MANUFACTURER = st.selectbox("MANUFACTURER", options=get_filter()['MANUFACTURER'].unique().tolist(), index=None)
        FLTR_ID = st.selectbox("Id", options=get_filter()['Id'].unique().tolist(), index=None)

        if FLTR_ID:
            TEMPLATE_ID = holder_template.selectbox("TEMPLATE Id", options=DATAFRAME['Id'].to_list(), index=DATAFRAME['Id'].to_list().index(FLTR_ID), label_visibility='collapsed')

if TEMPLATE_ID:
    SQL = SQL_BY_ROW("TEMPLATES", "Id", TEMPLATE_ID)
    # st.json(SQL)
    CURRENT_TEMPLATE = TEMPLATE.TypeDict(**SQL[0])
    
    ## DB DATA
    CURRENT_DB: dict = None
    if isinstance(CURRENT_TEMPLATE["DB"], str):
        try:
            CURRENT_DB = json.loads(CURRENT_TEMPLATE["DB"])
        except:
            CURRENT_DB = dict()
    elif isinstance(CURRENT_TEMPLATE["DB"], dict):
        CURRENT_DB = CURRENT_TEMPLATE["DB"]
    else:
        CURRENT_DB = dict()
    
    # st.write(CURRENT_DB)

    ## PRINT EXCEL
    ## __________________________________________________________________________________________________

    PRINT_XLSX(CURRENT_TEMPLATE)


    ## INFO
    ## __________________________________________________________________________________________________

    st.text("") # SEPARATOR
    # st.markdown(''':blue-background[💊 CMC:]''')
    st.subheader('INFO:', divider='blue')

    INFO_EDITOR("TEMPLATES", TEMPLATE_ID, CURRENT_TEMPLATE["INFO"])


    ## TEST LIST
    ## __________________________________________________________________________________________________

    st.text("") # SEPARATOR
    # st.markdown(''':blue-background[💊 CMC:]''')
    st.subheader('TEST LIST:', divider='blue')

    if not CURRENT_DB.get("TEST_LIST"):
        CURRENT_DB["TEST_LIST"] = []

    # st.write(CURRENT_DB["TEST_LIST"])
    selected = TEST_EDITOR(TEMPLATE_ID, CURRENT_DB)
    # st.write(selected)


    ## PYDATA
    ## __________________________________________________________________________________________________

    st.text("")
    st.text("")
    st.subheader('PYDATA:', divider='blue')

    # st.sidebar.markdown("""
    # [➡️ PYDATA](#pydata)
    # """, unsafe_allow_html=True)

    if not CURRENT_TEMPLATE["PYDATA"]:
        CURRENT_TEMPLATE["PYDATA"] = str()
    
    PYDATA_EDITOR("TEMPLATES", TEMPLATE_ID, CURRENT_TEMPLATE["PYDATA"])


    ## DB DATA JSON
    ## __________________________________________________________________________________________________

    st.text("") # SEPARATOR
    st.text("") # SEPARATOR
    # st.markdown(''':blue-background[💊 DB DATA:]''')
    st.subheader('JSON DB DATA:', divider='blue')
    
    DB_EDITOR("TEMPLATES", TEMPLATE_ID, CURRENT_TEMPLATE["DB"])



## Dataframe selections / https://docs.streamlit.io/develop/api-reference/data/st.dataframe
## __________________________________________________________________________________________________

# if "df" not in st.session_state:
#     st.session_state.df = pd.DataFrame(
#         np.random.randn(12, 5), columns=["a", "b", "c", "d", "e"]
#     )

# event = st.dataframe(
#     st.session_state.df,
#     key="data",
#     on_select="rerun",
#     selection_mode=["multi-row", "multi-column"],
# )

# event.selection
